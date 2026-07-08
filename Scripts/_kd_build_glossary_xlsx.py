"""Build KD-Translation-Glossary.xlsx from MD-Mapping and KD-Glossary-Additions.md."""
from __future__ import annotations

import argparse
import re
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from _common import BASE

MD_MAPPING = BASE / "References" / "Madhyasth-Darshan" / "MD-Mapping.xlsx"
GLOSSARY_MD = (
    BASE
    / "References"
    / "Madhyasth-Darshan"
    / "KD-Karm-Darshan-English"
    / "KD-Glossary-Additions.md"
)
OUTPUT = (
    BASE
    / "References"
    / "Madhyasth-Darshan"
    / "KD-Karm-Darshan-English"
    / "KD-Translation-Glossary.xlsx"
)

TABLE_ROW_RE = re.compile(
    r"^\|\s*(?P<hindi>[^|]+)\s*\|\s*(?P<translit>[^|]+)\s*\|\s*(?P<english>[^|]+)\s*\|\s*(?P<note>[^|]*)\s*\|$"
)


def parse_glossary_md(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    in_table = False
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("| Hindi |"):
            in_table = True
            continue
        if not in_table:
            continue
        if line.startswith("| :") or not line.startswith("|"):
            continue
        m = TABLE_ROW_RE.match(line.strip())
        if not m:
            continue
        rows.append(
            {
                "hindi": m.group("hindi").strip(),
                "transliteration": m.group("translit").strip(),
                "english": m.group("english").strip(),
                "note": m.group("note").strip(),
            }
        )
    return rows


def hindi_variants(hindi_cell: str) -> list[str]:
    return [v.strip() for v in hindi_cell.split(",") if v.strip()]


def load_md_mapping_lookup(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        hindi = str(row[0]).strip()
        english = row[1]
        if english is None:
            continue
        en_str = str(english).strip()
        for variant in hindi_variants(hindi):
            lookup.setdefault(variant, en_str)
    wb.close()
    return lookup


def build_overrides(
    md_lookup: dict[str, str],
    kd_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    overrides: list[dict[str, str]] = []
    seen: set[str] = set()
    for kd in kd_rows:
        for variant in hindi_variants(kd["hindi"]):
            if variant in seen:
                continue
            md_en = md_lookup.get(variant)
            if not md_en:
                continue
            kd_en = kd["english"].strip()
            if not kd_en or kd_en.lower() == md_en.lower():
                continue
            seen.add(variant)
            overrides.append(
                {
                    "hindi": variant,
                    "md_mapping_english": md_en,
                    "kd_english": kd_en,
                    "note": kd["note"],
                }
            )
    return overrides


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), max_width)


def build_workbook(out_path: Path) -> None:
    kd_rows = parse_glossary_md(GLOSSARY_MD)
    md_lookup = load_md_mapping_lookup(MD_MAPPING)
    override_rows = build_overrides(md_lookup, kd_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        shutil.copyfile(MD_MAPPING, tmp_path)
        wb = load_workbook(tmp_path)
        for name in list(wb.sheetnames):
            if name not in (wb.sheetnames[0],):
                del wb[name]
        first = wb.sheetnames[0]
        wb[first].title = "MD-Mapping"
        if "KD-Additions" in wb.sheetnames:
            del wb["KD-Additions"]
        if "KD-Overrides" in wb.sheetnames:
            del wb["KD-Overrides"]

        ws_kd = wb.create_sheet("KD-Additions")
        ws_kd.append(["Hindi", "Transliteration", "English used", "Note"])
        for row in kd_rows:
            ws_kd.append([row["hindi"], row["transliteration"], row["english"], row["note"]])

        ws_ov = wb.create_sheet("KD-Overrides")
        ws_ov.append(["Hindi", "MD-Mapping English", "KD English", "KD Note"])
        for row in override_rows:
            ws_ov.append(
                [row["hindi"], row["md_mapping_english"], row["kd_english"], row["note"]]
            )

        autosize_columns(ws_kd)
        autosize_columns(ws_ov)
        wb.save(out_path)
        wb.close()
    finally:
        tmp_path.unlink(missing_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build KD-Translation-Glossary.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT,
        help="Output xlsx path",
    )
    args = parser.parse_args()

    if not MD_MAPPING.is_file():
        print(f"Error: MD-Mapping not found: {MD_MAPPING}", file=sys.stderr)
        return 1
    if not GLOSSARY_MD.is_file():
        print(f"Error: glossary markdown not found: {GLOSSARY_MD}", file=sys.stderr)
        return 1

    build_workbook(args.output.resolve())
    kd_count = len(parse_glossary_md(GLOSSARY_MD))
    print(f"Wrote {args.output.resolve()} ({kd_count} KD-Additions rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
