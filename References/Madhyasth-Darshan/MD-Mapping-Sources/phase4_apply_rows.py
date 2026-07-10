"""Append Phase 4 proposed rows to MD-Mapping.xlsx starting at row 2148."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE.parent / "MD-Mapping.xlsx"
START_ROW = 2148


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--rows",
        type=Path,
        default=HERE / "phase4_new_rows.json",
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--start-row", type=int, default=START_ROW)
    args = parser.parse_args()

    proposals = json.loads(args.rows.read_text(encoding="utf-8"))
    wb = load_workbook(args.xlsx)
    ws = wb.active

    # Ensure we don't overwrite existing content.
    last_used = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            last_used = r
    start = max(args.start_row, last_used + 1)
    if start != args.start_row:
        print(f"note: starting at row {start} (last used {last_used})", file=sys.stderr)

    written = 0
    for i, row in enumerate(proposals):
        r = start + i
        ws.cell(r, 1).value = row["hindi_lemma"]
        ws.cell(r, 2).value = row["english"]
        ws.cell(r, 3).value = row.get("transliteration") or ""
        note = row.get("note") or "Phase 4 (freq≥2 exhaustive)"
        ws.cell(r, 6).value = note
        ws.cell(r, 9).value = row.get("citation") or ""
        written += 1

    wb.save(args.xlsx)
    print(
        f"wrote {written} rows to {args.xlsx} starting at {start} "
        f"(through {start + written - 1})",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
