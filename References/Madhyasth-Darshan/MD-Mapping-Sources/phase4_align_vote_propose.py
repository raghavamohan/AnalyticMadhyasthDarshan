"""Propose rows by positional alignment on short bilingual units.

When a pair has 2–5 Hindi tokens and 2–8 English content words, vote for
token→word alignments by index. High-confidence recurring alignments become
glossary proposals for uncovered tokens.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import EXTRA_STOP, clean_english, normalize_lemma, transliterate
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA = re.compile(r"[ऀ-ॿ]+")
LATIN = re.compile(r"[A-Za-z][A-Za-z'\-/]*")

EN_STOP = {
    "the",
    "a",
    "an",
    "and",
    "or",
    "of",
    "to",
    "in",
    "on",
    "for",
    "with",
    "from",
    "by",
    "as",
    "is",
    "are",
    "was",
    "were",
    "be",
    "been",
    "that",
    "this",
    "these",
    "those",
    "it",
    "its",
    "their",
    "his",
    "her",
    "who",
    "which",
    "what",
    "when",
    "where",
    "while",
    "through",
    "into",
    "also",
    "only",
    "more",
    "most",
    "such",
    "than",
    "then",
    "not",
    "no",
    "so",
    "if",
    "but",
    "can",
    "will",
    "have",
    "has",
    "had",
    "do",
    "does",
    "did",
    "one",
    "two",
    "three",
    "four",
    "five",
}


def load_known() -> tuple[set[str], set[str]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()
    return exact, stems


def content_words(text: str) -> list[str]:
    return [w for w in LATIN.findall(text) if w.lower() not in EN_STOP and len(w) > 2]


def main() -> None:
    exact, stems = load_known()
    stop = STOPWORDS | EXTRA_STOP
    votes: dict[str, Counter[str]] = defaultdict(Counter)
    cite: dict[tuple[str, str], str] = {}

    for name in ("mvd_pairs.json", "sb_pairs.json"):
        for p in json.loads((HERE / name).read_text(encoding="utf-8")):
            if not p.get("en"):
                continue
            hi_toks = DEVA.findall(p["hi"])
            en = clean_english(p["en"])
            en_words = content_words(en)
            if not (2 <= len(hi_toks) <= 5 and 2 <= len(en_words) <= 8):
                continue
            # Positional + endpoints alignment votes.
            n = min(len(hi_toks), len(en_words))
            for i in range(n):
                ht = hi_toks[i]
                ew = en_words[i]
                votes[ht][ew.lower()] += 2 if p["book"] == "MVD" else 1
                cite[(ht, ew.lower())] = f"{p['book']} p.{p['hi_page']}"
            # Also vote last-to-last (head-final Hindi often).
            votes[hi_toks[-1]][en_words[-1].lower()] += 1
            cite[(hi_toks[-1], en_words[-1].lower())] = f"{p['book']} p.{p['hi_page']}"

    freq_map = {
        tok: freq
        for tok, freq in json.loads((HERE / "candidates_freq2.json").read_text(encoding="utf-8"))
    }
    existing = json.loads((HERE / "phase4_new_rows.json").read_text(encoding="utf-8"))
    have = {r["token"] for r in existing}

    added = []
    for tok, counter in votes.items():
        if tok in have or tok in exact or tok in stop or len(tok) < 2:
            continue
        if freq_map.get(tok, 0) < 2:
            continue
        stem = normalize_lemma(tok)
        if stem in stems and stem != tok:
            continue
        gloss_l, n = counter.most_common(1)[0]
        total = sum(counter.values())
        if n < 3:
            continue
        if n / total < 0.35:
            continue
        # Prefer multi-letter non-generic glosses.
        if gloss_l in {
            "nature",
            "life",
            "form",
            "order",
            "state",
            "activity",
            "human",
            "humans",
            "study",
            "method",
            "process",
            "result",
            "effort",
            "motion",
            "world",
            "time",
            "space",
            "force",
            "power",
            "unit",
            "units",
            "earth",
            "family",
            "society",
            "tradition",
            "knowledge",
            "understanding",
            "resolution",
            "existence",
            "coexistence",
            "awakening",
            "development",
            "behaviour",
            "behavior",
            "evidence",
            "means",
            "used",
            "every",
            "also",
            "itself",
        }:
            continue
        english = gloss_l
        # Restore a citation casing from evidence if possible — title-case single word.
        english = english if english.islower() else english
        citation = cite.get((tok, gloss_l), "MVD")
        added.append(
            {
                "token": tok,
                "action": "propose",
                "hindi_lemma": tok,
                "english": english,
                "transliteration": transliterate(tok),
                "note": (
                    f"Phase 4 (freq≥2 exhaustive): English from positional alignment "
                    f"votes ({n}/{total}). Token freq={freq_map.get(tok)}."
                ),
                "citation": citation,
                "freq": freq_map.get(tok),
                "method": "align-vote",
                "vote_n": n,
                "vote_total": total,
            }
        )
        have.add(tok)

    merged = existing + added
    merged.sort(key=lambda r: (-(r.get("freq") or 0), r["hindi_lemma"]))
    (HERE / "phase4_new_rows.json").write_text(
        json.dumps(merged, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(
        f"existing={len(existing)} align_added={len(added)} total={len(merged)}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
