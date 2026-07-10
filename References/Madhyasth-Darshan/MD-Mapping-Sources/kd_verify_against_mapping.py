"""Compare KD working translation terminology against MD-Mapping.xlsx."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
BASE = HERE.parent
KD_DIR = BASE / "KD-Karm-Darshan-English"
DEFAULT_XLSX = BASE / "MD-Mapping.xlsx"
DEFAULT_KD_MD = KD_DIR / "KD-Karm-Darshan-English.md"
DEFAULT_KD_GLOSS = KD_DIR / "KD-Glossary-Additions.md"

TABLE_ROW_RE = re.compile(
    r"^\|\s*(?P<hindi>[^|]+)\s*\|\s*(?P<translit>[^|]+)\s*\|\s*(?P<english>[^|]+)\s*\|\s*(?P<note>[^|]*)\s*\|$"
)
ITALIC_RE = re.compile(r"(?<!\*)\*([^*]+)\*(?!\*)")
DEVA = re.compile(r"[ऀ-ॿ]+")

# Known deliberate KD exceptions (do not treat as accidental mismatches).
EXCEPTIONS = [
    {
        "hindi": "ताप",
        "md_english": "heat",
        "kd_english": "temperature",
        "reason": "KD 3.7 contrasts ताप=temperature with ऊष्मा=heat in the same passage.",
    },
    {
        "hindi": "वास्तविकता",
        "md_english": "reality",
        "kd_english": "actuality",
        "reason": "KD lists reality/actuality/truth as three distinct terms (यथार्थता/वास्तविकता/सत्य).",
    },
    {
        "hindi": "संस्कार",
        "md_english": None,
        "kd_english": "sanskar (transliterated)",
        "reason": "KD keeps संस्कार transliterated, distinct from संस्कृति=culture.",
    },
    {
        "hindi": "काम",
        "md_english": "lust",
        "kd_english": "kama (transliterated)",
        "reason": "KD keeps काम transliterated in moksha/dharma/kama/wealth list.",
    },
]


def load_md_mapping(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        hindi = str(row[0]).strip()
        english = str(row[1]).strip() if row[1] else ""
        translit = str(row[2]).strip() if row[2] else ""
        variants = [v.strip() for v in hindi.replace("،", ",").split(",") if v.strip()]
        rows.append(
            {
                "row": i,
                "hindi": hindi,
                "variants": variants,
                "english": english,
                "transliteration": translit,
            }
        )
    wb.close()
    return rows


def parse_kd_glossary(path: Path) -> list[dict]:
    rows: list[dict] = []
    in_table = False
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("| Hindi |"):
            in_table = True
            continue
        if not in_table:
            continue
        if line.startswith("| :") or not line.startswith("|"):
            continue
        m = TABLE_ROW_RE.match(line.strip())
        if not m:
            continue
        rows.append(
            {
                "hindi": m.group("hindi").strip(),
                "transliteration": m.group("translit").strip(),
                "english": m.group("english").strip(),
                "note": m.group("note").strip(),
            }
        )
    return rows


def norm_en(text: str) -> str:
    t = text.lower().strip()
    t = re.sub(r"\s+", " ", t)
    t = t.strip(" .;:,()/\"'")
    return t


def hindi_overlap(kd_hindi: str, md_variants: list[str]) -> bool:
    kd_toks = set(DEVA.findall(kd_hindi)) | {p.strip() for p in kd_hindi.split(",") if p.strip()}
    md_toks = set()
    for v in md_variants:
        md_toks.add(v)
        md_toks.update(DEVA.findall(v))
    return bool(kd_toks & md_toks)


def compare_glossary(kd_rows: list[dict], md_rows: list[dict]) -> list[dict]:
    out = []
    for kd in kd_rows:
        matches = [m for m in md_rows if hindi_overlap(kd["hindi"], m["variants"])]
        if not matches:
            out.append(
                {
                    "status": "md_missing",
                    "kd_hindi": kd["hindi"],
                    "kd_english": kd["english"],
                    "md_english": None,
                    "md_row": None,
                }
            )
            continue
        # Prefer exact variant match, else first overlap.
        md = matches[0]
        for m in matches:
            if any(v == kd["hindi"].split(",")[0].strip() for v in m["variants"]):
                md = m
                break
        kd_en = norm_en(kd["english"])
        md_en = norm_en(md["english"])
        if not md_en:
            status = "md_missing_english"
        elif kd_en == md_en or kd_en in md_en or md_en in kd_en:
            status = "aligned"
        else:
            status = "override"
        out.append(
            {
                "status": status,
                "kd_hindi": kd["hindi"],
                "kd_english": kd["english"],
                "md_english": md["english"],
                "md_row": md["row"],
            }
        )
    return out


def body_coverage(kd_text: str, md_rows: list[dict]) -> dict:
    text_l = kd_text.lower()
    present = []
    absent = []
    for m in md_rows:
        if not m["english"]:
            continue
        # Use primary English alternative before slash/paren.
        primary = re.split(r"[/;,(]", m["english"])[0].strip()
        if len(primary) < 4:
            continue
        if primary.lower() in text_l:
            present.append({"row": m["row"], "hindi": m["hindi"], "english": m["english"]})
        else:
            absent.append({"row": m["row"], "hindi": m["hindi"], "english": m["english"]})
    return {
        "english_glosses_checked": len(present) + len(absent),
        "present_in_kd_body": len(present),
        "absent_from_kd_body": len(absent),
        "present_sample": present[:40],
        "absent_sample_high_value": [
            a
            for a in absent
            if any(
                k in a["hindi"]
                for k in (
                    "सहअस्तित्व",
                    "जागृति",
                    "विकास",
                    "जीवन",
                    "मानव",
                    "अस्तित्व",
                    "समाधान",
                    "न्याय",
                    "धर्म",
                    "सत्य",
                )
            )
        ][:40],
    }


def italic_matches(kd_text: str, md_rows: list[dict]) -> list[dict]:
    italics = {m.group(1).strip().lower() for m in ITALIC_RE.finditer(kd_text)}
    hits = []
    for m in md_rows:
        t = (m["transliteration"] or "").lower().strip()
        if t and t in italics:
            hits.append(
                {
                    "italic": t,
                    "hindi": m["hindi"],
                    "english": m["english"],
                    "md_row": m["row"],
                }
            )
    return hits


def render_md(report: dict) -> str:
    g = report["glossary_comparison"]
    counts = report["glossary_status_counts"]
    cov = report["body_coverage"]
    lines = [
        "# KD vs MD-Mapping verification report",
        "",
        "Generated by `kd_verify_against_mapping.py`.",
        "",
        "## Summary",
        "",
        f"- MD-Mapping rows: **{report['md_mapping_rows']}**",
        f"- KD-Glossary-Additions rows: **{report['kd_glossary_rows']}**",
        f"- Glossary compare: aligned={counts.get('aligned', 0)}, "
        f"override={counts.get('override', 0)}, "
        f"md_missing={counts.get('md_missing', 0)}, "
        f"md_missing_english={counts.get('md_missing_english', 0)}",
        f"- KD body English gloss coverage: "
        f"{cov['present_in_kd_body']}/{cov['english_glosses_checked']} primary glosses found in body",
        f"- Italic transliteration matches: **{len(report['italic_matches'])}**",
        "",
        "## Known deliberate exceptions",
        "",
    ]
    for ex in report["exceptions"]:
        lines.append(
            f"- **{ex['hindi']}**: KD uses `{ex['kd_english']}`"
            + (f" vs MD `{ex['md_english']}`" if ex["md_english"] else "")
            + f" — {ex['reason']}"
        )
    lines += ["", "## KD-Glossary-Additions overrides (English differs)", ""]
    overrides = [x for x in g if x["status"] == "override"]
    if not overrides:
        lines.append("_None._")
    else:
        lines.append("| KD Hindi | KD English | MD-Mapping English | MD row |")
        lines.append("|---|---|---|---|")
        for x in overrides[:80]:
            lines.append(
                f"| {x['kd_hindi']} | {x['kd_english']} | {x['md_english']} | {x['md_row']} |"
            )
        if len(overrides) > 80:
            lines.append(f"| … | … | … | +{len(overrides) - 80} more |")
    lines += ["", "## KD-Glossary terms missing from MD-Mapping", ""]
    missing = [x for x in g if x["status"] == "md_missing"]
    if not missing:
        lines.append("_None._")
    else:
        for x in missing[:60]:
            lines.append(f"- `{x['kd_hindi']}` → {x['kd_english']}")
        if len(missing) > 60:
            lines.append(f"- … +{len(missing) - 60} more")
    lines += [
        "",
        "## High-value MD-Mapping glosses not found in KD body",
        "",
        "_Absence is often expected (term not yet used in translated chapters); not automatically an error._",
        "",
    ]
    for a in cov["absent_sample_high_value"][:30]:
        lines.append(f"- row {a['row']}: {a['hindi']} → {a['english']}")
    lines += ["", "## Aligned glossary sample", ""]
    aligned = [x for x in g if x["status"] == "aligned"][:20]
    for x in aligned:
        lines.append(f"- `{x['kd_hindi']}`: KD `{x['kd_english']}` ≈ MD `{x['md_english']}`")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--kd-md", type=Path, default=DEFAULT_KD_MD)
    parser.add_argument("--kd-glossary", type=Path, default=DEFAULT_KD_GLOSS)
    parser.add_argument(
        "--json-out",
        type=Path,
        default=HERE / "KD-vs-MD-Mapping-report.json",
    )
    parser.add_argument(
        "--md-out",
        type=Path,
        default=HERE / "KD-vs-MD-Mapping-report.md",
    )
    args = parser.parse_args()

    md_rows = load_md_mapping(args.xlsx)
    kd_gloss = parse_kd_glossary(args.kd_glossary)
    kd_text = args.kd_md.read_text(encoding="utf-8")
    gloss_cmp = compare_glossary(kd_gloss, md_rows)
    status_counts: dict[str, int] = {}
    for item in gloss_cmp:
        status_counts[item["status"]] = status_counts.get(item["status"], 0) + 1

    report = {
        "md_mapping_rows": len(md_rows),
        "kd_glossary_rows": len(kd_gloss),
        "glossary_comparison": gloss_cmp,
        "glossary_status_counts": status_counts,
        "exceptions": EXCEPTIONS,
        "body_coverage": body_coverage(kd_text, md_rows),
        "italic_matches": italic_matches(kd_text, md_rows),
    }
    args.json_out.write_text(
        json.dumps(report, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    args.md_out.write_text(render_md(report), encoding="utf-8", newline="\n")
    print(
        f"md_rows={len(md_rows)} kd_gloss={len(kd_gloss)} "
        f"status={status_counts} -> {args.md_out}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
