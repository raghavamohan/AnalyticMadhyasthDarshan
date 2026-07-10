"""Augment phase4_new_rows.json using shortest focused evidence units."""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import (
    EXTRA_STOP,
    clean_english,
    first_sentence,
    normalize_lemma,
    transliterate,
)
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA = re.compile(r"[ऀ-ॿ]+")
LATIN = re.compile(r"[A-Za-z][A-Za-z'\-/]*")
BAD = {
    "जब",
    "तब",
    "बात",
    "ने",
    "जा",
    "सब",
    "आज",
    "अत",
    "यातु",
    "कम",
    "बन",
    "नाम",
    "अध्याय",
    "दो",
    "तीन",
    "चार",
    "पाँच",
    "पांच",
    "एक",
    "नागराज",
}


def load_known() -> tuple[set[str], set[str]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()
    return exact, stems


def main() -> None:
    path = HERE / "phase4_new_rows.json"
    existing = json.loads(path.read_text(encoding="utf-8"))
    have = {r["token"] for r in existing}
    exact, stems = load_known()
    stop = STOPWORDS | EXTRA_STOP | BAD
    evidence = json.loads((HERE / "phase4_evidence.json").read_text(encoding="utf-8"))
    added: list[dict] = []

    for item in evidence:
        tok = item["token"]
        if tok in have or tok in stop or tok in exact or len(tok) < 2:
            continue
        stem = normalize_lemma(tok)
        if stem in stems and stem != tok:
            continue
        hits: list[tuple] = []
        for h in item["evidence"]:
            en = clean_english(h.get("en") or "")
            if not en:
                continue
            hi_toks = DEVA.findall(h["hi"])
            hi_n = len(hi_toks)
            if hi_n > 6 or tok not in hi_toks:
                continue
            gloss = first_sentence(en, max_words=6)
            if not gloss:
                continue
            wc = len(LATIN.findall(gloss))
            if not (1 <= wc <= 6):
                continue
            low = gloss.lower()
            if low.startswith(
                ("chapter", "may ", "according", "every ", "when ", "the activity", "buddhi")
            ):
                continue
            if low in {"jeevan", "big", "world", "mirage", "kindness", "conceit"}:
                continue
            hits.append((hi_n, 0 if h["book"] == "MVD" else 1, wc, h, gloss))
        if not hits:
            continue
        hits.sort()
        hi_n, _, wc, h, gloss = hits[0]
        if hi_n > 3 and wc > 4:
            continue
        added.append(
            {
                "token": tok,
                "action": "propose",
                "hindi_lemma": tok,
                "english": gloss,
                "transliteration": transliterate(tok),
                "note": (
                    "Phase 4 (freq≥2 exhaustive): English from shortest focused "
                    f"evidence unit ({h['book']}). Token freq={item['total_occurrences']}."
                ),
                "citation": f"{h['book']} p.{h['hi_page']}",
                "freq": item["total_occurrences"],
                "method": "evidence-short",
                "evidence_book": h["book"],
                "evidence_page": h["hi_page"],
            }
        )
        have.add(tok)

    merged = existing + added
    merged.sort(key=lambda r: (-(r.get("freq") or 0), r["hindi_lemma"]))
    path.write_text(json.dumps(merged, ensure_ascii=False, indent=1) + "\n", encoding="utf-8", newline="\n")
    print(
        f"existing={len(existing)} added={len(added)} total={len(merged)} "
        f"methods={Counter(r.get('method') for r in merged)}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
