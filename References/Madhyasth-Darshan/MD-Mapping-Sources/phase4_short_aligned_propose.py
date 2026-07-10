"""Build Phase 4 rows from short, well-aligned MVD/SB hi→en units.

Only uses pairs where Hindi is short (1–6 Devanagari tokens) and English is a
clean short gloss (1–8 Latin words). This favors glossary-like bilingual items
over narrative paragraphs.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import (
    EXTRA_STOP,
    clean_english,
    first_sentence,
    normalize_lemma,
    transliterate,
    jv_pages_for_english,
)
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA = re.compile(r"[ऀ-ॿ]+")
LATIN = re.compile(r"[A-Za-z][A-Za-z'\-/]*")
QUOTED = re.compile(r"[\"'‘’“”]([^\"'‘’“”]{2,50})[\"'‘’“”]")

NUMBER_WORDS = {
    "एक",
    "दो",
    "तीन",
    "चार",
    "पाँच",
    "पांच",
    "छह",
    "सात",
    "आठ",
    "नौ",
    "दस",
}
BAD_TOKENS = {
    "जब",
    "तब",
    "बात",
    "नागराज",
    "अध्याय",
    "ने",
    "जा",
    "सब",
    "आज",
    "अत",
    "यातु",
    "कम",
    "बन",
    "बनी",
    "वश",
    "नाम",
}


def load_known(xlsx: Path) -> tuple[set[str], set[str]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()
    return exact, stems


def load_pairs() -> list[dict]:
    out = []
    for name in ("mvd_pairs.json", "sb_pairs.json"):
        data = json.loads((HERE / name).read_text(encoding="utf-8"))
        out.extend(data)
    return out


def gloss_from_pair(hi: str, en: str) -> str | None:
    en_c = clean_english(en)
    if not en_c:
        return None
    for q in QUOTED.findall(en_c):
        q = q.strip()
        words = LATIN.findall(q)
        if 1 <= len(words) <= 6:
            return q
    gloss = first_sentence(en_c, max_words=8)
    if not gloss:
        return None
    words = LATIN.findall(gloss)
    if not (1 <= len(words) <= 8):
        return None
    low = gloss.lower()
    if low.startswith(("chapter", "page", "http", "www", "may ")):
        return None
    return gloss


def main() -> None:
    exact, stems = load_known(XLSX)
    stop = STOPWORDS | EXTRA_STOP | NUMBER_WORDS | BAD_TOKENS
    pairs = load_pairs()
    jv_path = HERE.parent / "JV-Jeevan-Vidya-An-Introduction.md"
    jv_text = jv_path.read_text(encoding="utf-8") if jv_path.is_file() else ""

    # token -> best proposal
    best: dict[str, dict] = {}

    for p in pairs:
        if not p.get("en"):
            continue
        hi = p["hi"]
        en = p["en"]
        toks = DEVA.findall(hi)
        if not (1 <= len(toks) <= 6):
            continue
        gloss = gloss_from_pair(hi, en)
        if not gloss:
            continue
        # Prefer single-token Hindi units; for multi-token, only add if unit is
        # definitional (contains संज्ञा / तात्पर्य / :- ) or exactly 1–2 content tokens.
        definitional = bool(re.search(r"(संज्ञा|तात्पर्य|अर्थ)|:-", hi))
        if len(toks) > 2 and not definitional:
            continue

        citation = f"{p['book']} p.{p['hi_page']}"
        for token in toks:
            if token in stop or len(token) < 2:
                continue
            if token in exact:
                continue
            stem = normalize_lemma(token)
            if stem in stems and stem != token:
                continue
            # For multi-token Hindi, only map the "head" content tokens (skip if
            # gloss is clearly about another known term — still keep candidate).
            score = (0 if p["book"] == "MVD" else 1, len(toks), len(LATIN.findall(gloss)))
            row = {
                "token": token,
                "action": "propose",
                "hindi_lemma": token,
                "english": gloss,
                "transliteration": transliterate(token),
                "note": (
                    "Phase 4 (freq≥2 exhaustive): English from short aligned "
                    f"{p['book']} bilingual unit ({len(toks)} hi tokens)."
                ),
                "citation": citation,
                "freq": None,
                "method": "short-aligned",
                "evidence_book": p["book"],
                "evidence_page": p["hi_page"],
                "_score": score,
            }
            prev = best.get(token)
            if prev is None or score < tuple(prev["_score"]):
                best[token] = row

    # Attach frequencies from candidates file if present.
    freq_map = {}
    cand_path = HERE / "candidates_freq2.json"
    if cand_path.is_file():
        for tok, freq in json.loads(cand_path.read_text(encoding="utf-8")):
            freq_map[tok] = freq

    # Also merge high-confidence filtered seed rows.
    seed_path = HERE / "phase4_new_rows_filtered.json"
    seed = []
    if seed_path.is_file():
        seed = json.loads(seed_path.read_text(encoding="utf-8"))

    proposals: list[dict] = []
    seen = set()
    for row in seed:
        tok = row["token"]
        if tok in stop:
            continue
        row = dict(row)
        row["freq"] = freq_map.get(tok, row.get("freq"))
        # JV citation enrichment
        pages = jv_pages_for_english(row["english"], jv_text)
        if pages and "JV p." not in row["citation"]:
            row["citation"] = row["citation"] + "; " + "; ".join(f"JV p.{p}" for p in pages)
        proposals.append(row)
        seen.add(tok)

    for tok, row in best.items():
        if tok in seen:
            continue
        if freq_map and tok not in freq_map:
            # Only add tokens that appear at least twice in the corpus overall.
            continue
        if freq_map.get(tok, 0) < 2:
            continue
        row = dict(row)
        row["freq"] = freq_map.get(tok)
        row.pop("_score", None)
        pages = jv_pages_for_english(row["english"], jv_text)
        if pages:
            row["citation"] = row["citation"] + "; " + "; ".join(f"JV p.{p}" for p in pages)
        proposals.append(row)
        seen.add(tok)

    # Drop remaining obvious junk.
    cleaned = []
    for row in proposals:
        en = row["english"].strip()
        words = LATIN.findall(en)
        if not words:
            continue
        if len(words) > 8:
            continue
        low = en.lower()
        if low in {"jeevan", "big", "world", "mirage", "kindness", "conceit"}:
            continue
        if row["token"] in BAD_TOKENS | NUMBER_WORDS:
            continue
        cleaned.append(row)

    cleaned.sort(key=lambda r: (-(r.get("freq") or 0), r["hindi_lemma"]))
    out = HERE / "phase4_new_rows.json"
    out.write_text(json.dumps(cleaned, ensure_ascii=False, indent=1) + "\n", encoding="utf-8", newline="\n")
    print(f"proposals={len(cleaned)} (seed={len(seed)}) -> {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
