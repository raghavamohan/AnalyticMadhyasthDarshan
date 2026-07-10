"""Build uncovered Hindi token candidates from MVD/SB pairs vs MD-Mapping.xlsx."""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from stopwords import STOPWORDS

DEVA_TOKEN = re.compile(r"[ऀ-ॿ]+")
HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE.parent / "MD-Mapping.xlsx"


def load_known_tokens(xlsx: Path) -> set[str]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    known: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            known.add(cell)
            known.update(DEVA_TOKEN.findall(cell))
    wb.close()
    return known


def load_pair_token_counts(paths: list[Path]) -> Counter[str]:
    ctr: Counter[str] = Counter()
    for path in paths:
        data = json.loads(path.read_text(encoding="utf-8"))
        for pair in data:
            if not pair.get("en"):
                continue
            ctr.update(DEVA_TOKEN.findall(pair["hi"]))
    return ctr


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to MD-Mapping.xlsx",
    )
    parser.add_argument(
        "--min-freq",
        type=int,
        default=2,
        help="Minimum token frequency in pairs (default: 2)",
    )
    parser.add_argument(
        "--pairs",
        nargs="+",
        type=Path,
        default=[HERE / "mvd_pairs.json", HERE / "sb_pairs.json"],
        help="Pair JSON files",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=HERE / "candidates_freq2.json",
        help="Output [[token, freq], ...] JSON",
    )
    args = parser.parse_args()

    known = load_known_tokens(args.xlsx)
    counts = load_pair_token_counts(args.pairs)
    candidates = sorted(
        (
            (token, freq)
            for token, freq in counts.items()
            if freq >= args.min_freq
            and token not in known
            and token not in STOPWORDS
            and len(token) >= 2
        ),
        key=lambda item: (-item[1], item[0]),
    )
    args.output.write_text(
        json.dumps(candidates, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(
        f"known={len(known)} pair_tokens={len(counts)} "
        f"candidates_freq>={args.min_freq}={len(candidates)} -> {args.output}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
