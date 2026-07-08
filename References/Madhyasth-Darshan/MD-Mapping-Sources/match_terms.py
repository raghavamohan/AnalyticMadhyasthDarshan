import json, re, sys
import openpyxl

DEVA_TOKEN = re.compile(r'[ऀ-ॿ]+')

def tokenize(text):
    return DEVA_TOKEN.findall(text)

def load_pairs(path):
    data = json.load(open(path, encoding='utf-8'))
    out = []
    for p in data:
        if not p.get('en'):
            continue
        toks = tokenize(p['hi'])
        out.append({
            'hi_tokens': toks,
            'hi_text': p['hi'],
            'en_text': p['en'],
            'hi_page': p['hi_page'],
            'en_page': p['en_page'],
            'book': p['book'],
        })
    return out

def find_matches(term_tokens, pairs):
    n = len(term_tokens)
    hits = []
    for p in pairs:
        toks = p['hi_tokens']
        if n == 1:
            if term_tokens[0] in toks:
                hits.append(p)
        else:
            for i in range(len(toks) - n + 1):
                if toks[i:i+n] == term_tokens:
                    hits.append(p)
                    break
    return hits

def main():
    wb = openpyxl.load_workbook('MD-Mapping.xlsx', data_only=True)
    ws = wb['Sheet1']

    mvd_pairs = load_pairs(sys.argv[1])
    sb_pairs = load_pairs(sys.argv[2])
    out_path = sys.argv[3]

    print(f'MVD usable pairs: {len(mvd_pairs)}, SB usable pairs: {len(sb_pairs)}', file=sys.stderr)

    results = []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a or not str(a).strip():
            continue
        existing_en = ws.cell(r, 2).value
        existing_translit = ws.cell(r, 3).value
        existing_ref = ws.cell(r, 9).value
        variants = [t.strip() for t in str(a).split(',') if t.strip()]

        row_evidence = {'MVD': [], 'SB': []}
        for term in variants:
            term_tokens = tokenize(term)
            if not term_tokens:
                continue
            mvd_hits = find_matches(term_tokens, mvd_pairs)
            for h in mvd_hits:
                row_evidence['MVD'].append({
                    'term': term, 'hi_page': h['hi_page'], 'en_page': h['en_page'],
                    'hi_text': h['hi_text'], 'en_text': h['en_text'],
                })
            sb_hits = find_matches(term_tokens, sb_pairs)
            for h in sb_hits:
                row_evidence['SB'].append({
                    'term': term, 'hi_page': h['hi_page'], 'en_page': h['en_page'],
                    'hi_text': h['hi_text'], 'en_text': h['en_text'],
                })

        if not row_evidence['MVD'] and not row_evidence['SB']:
            continue  # no evidence at all — skip, nothing to reconcile

        results.append({
            'row': r,
            'hindi_variants': variants,
            'existing_english': existing_en,
            'existing_transliteration': existing_translit,
            'existing_reference': existing_ref,
            'evidence': row_evidence,
        })

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    with_mvd = sum(1 for x in results if x['evidence']['MVD'])
    with_sb_only = sum(1 for x in results if not x['evidence']['MVD'] and x['evidence']['SB'])
    print(f'Rows with any evidence: {len(results)}', file=sys.stderr)
    print(f'  with MVD evidence: {with_mvd}', file=sys.stderr)
    print(f'  SB-only evidence: {with_sb_only}', file=sys.stderr)

if __name__ == '__main__':
    main()
