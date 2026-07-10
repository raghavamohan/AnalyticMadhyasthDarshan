"""Write review cards for remaining technical candidates."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import EXTRA_STOP, normalize_lemma
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA = re.compile(r"[ऀ-ॿ]+")


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()

    have = {
        r["token"]
        for r in json.loads((HERE / "phase4_new_rows.json").read_text(encoding="utf-8"))
    }
    stop = STOPWORDS | EXTRA_STOP
    evidence = json.loads((HERE / "phase4_evidence.json").read_text(encoding="utf-8"))
    cards = []
    for item in evidence:
        tok = item["token"]
        if len(tok) < 4 or item["total_occurrences"] < 3:
            continue
        if tok in have or tok in exact or tok in stop:
            continue
        stem = normalize_lemma(tok)
        if stem in stems and stem != tok:
            continue
        if tok.endswith(("कर", "करने", "किया", "हुआ", "गया", "रहा")):
            continue
        e0 = item["evidence"][0]
        cards.append(
            {
                "token": tok,
                "freq": item["total_occurrences"],
                "hi": e0["hi"][:280].replace("\n", " "),
                "en": e0["en"][:280].replace("\n", " "),
                "book": e0["book"],
                "page": e0["hi_page"],
            }
        )
    cards.sort(key=lambda c: -c["freq"])
    (HERE / "phase4_review_cards.json").write_text(
        json.dumps(cards, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    lines = []
    for c in cards[:150]:
        lines.append(f"TOKEN={c['token']} FREQ={c['freq']} {c['book']} p.{c['page']}")
        lines.append(f"HI: {c['hi']}")
        lines.append(f"EN: {c['en']}")
        lines.append("")
    (HERE / "phase4_review_cards.txt").write_text("\n".join(lines), encoding="utf-8")
    print(f"cards={len(cards)} wrote review files")


if __name__ == "__main__":
    main()
