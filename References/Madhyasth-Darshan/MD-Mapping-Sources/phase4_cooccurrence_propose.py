"""Propose glossary rows by English co-occurrence across evidence hits.

For each Hindi candidate, collect Latin content-words from all paired English
snippets; if a content word/phrase recurs, treat it as the gloss. This is more
robust than taking a single misaligned short-unit sentence.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import (
    EXTRA_STOP,
    clean_english,
    normalize_lemma,
    transliterate,
)
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA_TOKEN = re.compile(r"[ऀ-ॿ]+")
LATIN_WORD = re.compile(r"[A-Za-z][A-Za-z'\-/]*")
QUOTED = re.compile(r"[\"'‘’“”]([^\"'‘’“”]{2,40})[\"'‘’“”]")

EN_STOP = {
    "the",
    "a",
    "an",
    "and",
    "or",
    "of",
    "to",
    "in",
    "on",
    "for",
    "with",
    "from",
    "by",
    "as",
    "is",
    "are",
    "was",
    "were",
    "be",
    "been",
    "being",
    "that",
    "this",
    "these",
    "those",
    "it",
    "its",
    "their",
    "his",
    "her",
    "our",
    "your",
    "who",
    "which",
    "what",
    "when",
    "where",
    "while",
    "through",
    "into",
    "over",
    "after",
    "before",
    "between",
    "among",
    "also",
    "only",
    "more",
    "most",
    "such",
    "than",
    "then",
    "there",
    "here",
    "not",
    "no",
    "nor",
    "so",
    "if",
    "but",
    "because",
    "about",
    "above",
    "below",
    "up",
    "down",
    "out",
    "off",
    "again",
    "further",
    "once",
    "all",
    "any",
    "both",
    "each",
    "few",
    "other",
    "some",
    "can",
    "will",
    "just",
    "should",
    "would",
    "could",
    "may",
    "might",
    "must",
    "shall",
    "have",
    "has",
    "had",
    "do",
    "does",
    "did",
    "human",
    "humans",
    "being",
    "beings",
    "one",
    "two",
    "three",
    "four",
    "five",
    "chapter",
    "page",
    "etc",
}


def load_known(xlsx: Path) -> tuple[set[str], set[str]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA_TOKEN.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA_TOKEN.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()
    return exact, stems


def content_words(text: str) -> list[str]:
    return [w.lower() for w in LATIN_WORD.findall(text) if w.lower() not in EN_STOP and len(w) > 2]


def bigrams(words: list[str]) -> list[str]:
    return [f"{words[i]} {words[i+1]}" for i in range(len(words) - 1)]


def propose_from_cooccurrence(item: dict) -> dict | None:
    token = item["token"]
    word_counts: Counter[str] = Counter()
    phrase_counts: Counter[str] = Counter()
    quote_counts: Counter[str] = Counter()
    best_hit = None
    best_score: tuple[int, int] | None = None

    for hit in item.get("evidence", []):
        en = clean_english(hit.get("en") or "")
        if not en:
            continue
        words = content_words(en)
        word_counts.update(words)
        phrase_counts.update(bigrams(words))
        for q in QUOTED.findall(en):
            q = q.strip()
            if 2 <= len(q) <= 40 and LATIN_WORD.search(q):
                quote_counts[q] += 1
        score = (0 if hit["book"] == "MVD" else 1, len(DEVA_TOKEN.findall(hit["hi"])) + len(words))
        if best_score is None or score < best_score:
            best_score = score
            best_hit = hit

    if not best_hit:
        return None

    english = None
    method = None
    if quote_counts:
        english, n = quote_counts.most_common(1)[0]
        if n >= 1 and len(content_words(english)) <= 5:
            method = "cooc-quoted"
    if not english and phrase_counts:
        phrase, n = phrase_counts.most_common(1)[0]
        # Require recurrence for phrases unless freq is high and phrase is technical-looking.
        if n >= 2 or (item["total_occurrences"] >= 5 and n >= 1):
            english = phrase.title() if phrase.islower() else phrase
            method = "cooc-bigram"
    if not english and word_counts:
        word, n = word_counts.most_common(1)[0]
        if n >= 2 and word not in {"nature", "life", "form", "order", "state", "activity"}:
            english = word
            method = "cooc-word"
        elif n >= 3:
            english = word
            method = "cooc-word-strong"

    if not english or not method:
        return None

    citation = f"{best_hit['book']} p.{best_hit['hi_page']}"
    return {
        "token": token,
        "action": "propose",
        "hindi_lemma": token,
        "english": english if not english.islower() or " " in english else english,
        "transliteration": transliterate(token),
        "note": (
            f"Phase 4 (freq≥2 exhaustive): English from co-occurrence across paired "
            f"evidence ({method}). Token freq={item['total_occurrences']}."
        ),
        "citation": citation,
        "freq": item["total_occurrences"],
        "method": method,
        "evidence_book": best_hit["book"],
        "evidence_page": best_hit["hi_page"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--evidence", type=Path, default=HERE / "phase4_evidence.json")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument(
        "--seed",
        type=Path,
        default=HERE / "phase4_new_rows_filtered.json",
        help="Already-accepted high-confidence rows to merge",
    )
    parser.add_argument("-o", type=Path, default=HERE / "phase4_new_rows.json")
    args = parser.parse_args()

    evidence = json.loads(args.evidence.read_text(encoding="utf-8"))
    exact, stems = load_known(args.xlsx)
    stop = STOPWORDS | EXTRA_STOP
    seed = []
    if args.seed.is_file():
        seed = json.loads(args.seed.read_text(encoding="utf-8"))

    have = {r["token"] for r in seed}
    for r in seed:
        for part in r["hindi_lemma"].split(","):
            have.add(part.strip())

    added: list[dict] = []
    for item in evidence:
        token = item["token"]
        if token in stop or token in exact or token in have:
            continue
        stem = normalize_lemma(token)
        if stem in stems and stem != token:
            continue
        if item["total_occurrences"] < 2:
            continue
        row = propose_from_cooccurrence(item)
        if not row:
            continue
        # Skip ultra-generic single words unless frequent and specific-looking.
        en = row["english"]
        if " " not in en and en.lower() in {
            "nature",
            "life",
            "form",
            "order",
            "state",
            "activity",
            "human",
            "study",
            "evidence",
            "behaviour",
            "behavior",
            "development",
            "awakening",
            "existence",
            "coexistence",
            "resolution",
            "understanding",
            "knowledge",
            "tradition",
            "society",
            "family",
            "earth",
            "world",
            "time",
            "space",
            "force",
            "power",
            "unit",
            "units",
            "method",
            "process",
            "result",
            "effort",
            "motion",
        }:
            continue
        added.append(row)
        have.add(token)

    # Merge seed + added; prefer seed on token clash.
    merged = list(seed)
    seed_tokens = {r["token"] for r in seed}
    for row in added:
        if row["token"] not in seed_tokens:
            merged.append(row)

    merged.sort(key=lambda r: (-r.get("freq", 0), r["hindi_lemma"]))
    args.o.write_text(json.dumps(merged, ensure_ascii=False, indent=1) + "\n", encoding="utf-8", newline="\n")
    print(f"seed={len(seed)} cooc_added={len(added)} total={len(merged)} -> {args.o}", file=sys.stderr)


if __name__ == "__main__":
    main()
