"""Propose Phase 4 MD-Mapping rows from phase4_evidence.json.

Rejects function words, inflections of existing lemmas, and low-confidence
evidence. English is taken only from paired MVD/SB text (never invented).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
DEFAULT_XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA_TOKEN = re.compile(r"[ऀ-ॿ]+")
LATIN_WORD = re.compile(r"[A-Za-z][A-Za-z'\-/]*")
QUOTED = re.compile(r"[\"'‘’“”]([^\"'‘’“”]{2,60})[\"'‘’“”]")
PAGE_BOILER = re.compile(
    r"^(page[-\s]?\d+|madhyasth|co-?existentialism|resolution centred|"
    r"samadhanatmak|jee?van vidya|www\.|http)",
    re.I,
)

# Extra function / discourse words that dominate freq≥2 but are not glossary terms.
EXTRA_STOP = set(
    """
    प्रकार प्रत्येक हर कर ओर अनेक समस्त अधिक स्पष्ट संपूर्ण तात्पर्य समझ संपन्न
    धरती प्रमाणित जागृत संज्ञा स्वयं सभी कुछ बहुत इसी इसीलिए इसलिए जबकि
    अथवा तथा एवं या और किन्तु परन्तु अगर चाहे भी तो ही में से को का की के
    है हैं हो था थी थे हुआ हुई हुए गया गयी गए करते करता करती करने किया
    किये किए जाता जाती जाने जाना रहना रहता रहती रहते रहा रही रहे पाना
    पाता पाती पाया पाये पाई सकता सकती सकते होगा होगी होंगे वाला वाली वाले
    अपना अपनी अपने इसका इसकी इसके उसका उसकी उसके यह वह ये वे हम तुम मैं
    मुझे उन्हें इन उन जो कि पर तक भर हेतु द्वारा अनुसार क्योंकि जबकि
    अर्थात् अर्थात् अर्थात् अर्थात् अर्थात् अर्थात् अर्थात् अर्थात्
    इसीलिए इसलिए इसलिए इसलिए इसलिए इसलिए इसलिए इसलिए इसलिए
    उपरान्त उपरान्त उपरान्त उपरान्त उपरान्त उपरान्त उपरान्त
    फलस्वरुप फलस्वरूप प्रधानत मूलत केवल मात्र आदि बीच बिना सही लोग
    आदमी रुप जहाँ दूसरा अन्य अन्यथा उक्त वैसा कहाँ किंतु परंतु
    इन्हीं उन्हीं जिन बड़े लगता कहते उनके कैसे वहाँ बाद स्वयम्
    इसीलिये यथा समझना स्थितियाँ जिन्हें निम्नानुसार
    प्रकारसे प्रकारकी प्रकारका प्रकारके प्रकारों प्रकारोंमें
    हरएक हरएकमें हरएककी
    """.split()
)

COMMON_SUFFIXES = (
    "पूर्णता",
    "पूर्ण",
    "शीलता",
    "शील",
    "वादी",
    "वाद",
    "कर्ता",
    "कर्ता",
    "पन",
    "ता",
    "त्व",
    "ओं",
    "यों",
    "ें",
    "ों",
    "ाँ",
    "ां",
    "ीं",
    "ा",
    "ी",
    "े",
    "ो",
    "ु",
    "ू",
    "ं",
    "ः",
    "्",
)

# Simple Devanagari → Latin for column C (approximate, consistent with Phase 3 style).
_TRANS = {
    "अ": "a",
    "आ": "a",
    "इ": "i",
    "ई": "i",
    "उ": "u",
    "ऊ": "u",
    "ऋ": "ri",
    "ए": "e",
    "ऐ": "ai",
    "ओ": "o",
    "औ": "au",
    "क": "k",
    "ख": "kh",
    "ग": "g",
    "घ": "gh",
    "ङ": "ng",
    "च": "ch",
    "छ": "chh",
    "ज": "j",
    "झ": "jh",
    "ञ": "ny",
    "ट": "t",
    "ठ": "th",
    "ड": "d",
    "ढ": "dh",
    "ण": "n",
    "त": "t",
    "थ": "th",
    "द": "d",
    "ध": "dh",
    "न": "n",
    "प": "p",
    "फ": "ph",
    "ब": "b",
    "भ": "bh",
    "म": "m",
    "य": "y",
    "र": "r",
    "ल": "l",
    "व": "v",
    "श": "sh",
    "ष": "sh",
    "स": "s",
    "ह": "h",
    "क्ष": "ksh",
    "त्र": "tr",
    "ज्ञ": "gy",
    "ा": "a",
    "ि": "i",
    "ी": "i",
    "ु": "u",
    "ू": "u",
    "ृ": "ri",
    "े": "e",
    "ै": "ai",
    "ो": "o",
    "ौ": "au",
    "ं": "n",
    "ः": "h",
    "्": "",
    "ँ": "n",
    "़": "",
}


def transliterate(text: str) -> str:
    out: list[str] = []
    i = 0
    while i < len(text):
        if text[i : i + 2] in _TRANS:
            out.append(_TRANS[text[i : i + 2]])
            i += 2
            continue
        out.append(_TRANS.get(text[i], text[i] if not ("ऀ" <= text[i] <= "ॿ") else ""))
        i += 1
    s = "".join(out)
    s = re.sub(r"[^a-zA-Z\-]+", "", s)
    return s.lower() or "TODO"


def normalize_lemma(token: str) -> str:
    t = unicodedata.normalize("NFC", token.strip())
    changed = True
    while changed and len(t) > 2:
        changed = False
        for suf in COMMON_SUFFIXES:
            if t.endswith(suf) and len(t) - len(suf) >= 2:
                t = t[: -len(suf)]
                changed = True
                break
    return t


def load_known(xlsx: Path) -> tuple[set[str], set[str], dict[str, str]]:
    """Return exact tokens, normalized stems, and hindi→english lookup."""
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    lookup: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        en = str(row[1]).strip() if row[1] else ""
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA_TOKEN.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA_TOKEN.findall(cell):
                stems.add(normalize_lemma(tok))
                if en:
                    lookup.setdefault(tok, en)
            if en:
                lookup.setdefault(cell, en)
    wb.close()
    return exact, stems, lookup


def clean_english(text: str) -> str:
    # Drop leading Devanagari leftovers from mis-paired blocks.
    lines = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if DEVA_TOKEN.search(stripped) and not LATIN_WORD.search(stripped):
            continue
        # Strip leading Devanagari run before Latin starts.
        m = re.search(r"[A-Za-z]", stripped)
        if m and m.start() > 0 and DEVA_TOKEN.search(stripped[: m.start()]):
            stripped = stripped[m.start() :]
        if PAGE_BOILER.match(stripped):
            continue
        lines.append(stripped)
    return " ".join(lines).strip()


def first_sentence(text: str, max_words: int = 12) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return ""
    # Split on sentence end, keep first useful clause.
    parts = re.split(r"(?<=[.:;])\s+", text)
    for part in parts:
        part = part.strip(" .;:-")
        if len(part) < 3:
            continue
        if PAGE_BOILER.match(part):
            continue
        words = LATIN_WORD.findall(part)
        if not words:
            continue
        if len(words) > max_words:
            return " ".join(words[:max_words])
        # Prefer the original casing/punctuation of a short clause.
        if len(words) <= max_words:
            return part[:120].strip(" .;:-")
    words = LATIN_WORD.findall(text)
    return " ".join(words[:max_words]) if words else ""


def extract_gloss(token: str, hi: str, en: str) -> tuple[str, str] | None:
    """Return (english, method) or None."""
    en_clean = clean_english(en)
    if not en_clean or len(LATIN_WORD.findall(en_clean)) < 1:
        return None

    hi_tokens = DEVA_TOKEN.findall(hi)
    # Quoted English often carries the technical gloss.
    for q in QUOTED.findall(en_clean):
        if LATIN_WORD.search(q) and not DEVA_TOKEN.search(q):
            q = q.strip()
            if 2 <= len(q) <= 60:
                return q, "quoted"

    # Definitional short Hindi units dominated by the token.
    if token in hi_tokens and len(hi_tokens) <= 8:
        gloss = first_sentence(en_clean, max_words=10)
        if gloss and len(LATIN_WORD.findall(gloss)) >= 1:
            # Avoid dumping long narrative as a "gloss".
            if len(LATIN_WORD.findall(gloss)) <= 10:
                return gloss, "short-unit"

    # Pattern: "X means/implies/is called/referred to as Y"
    m = re.search(
        r"(?:means|implies|is called|referred to as|termed|denotes|named)\s+(.+?)(?:[.:;]|$)",
        en_clean,
        re.I,
    )
    if m:
        gloss = first_sentence(m.group(1), max_words=8)
        if gloss:
            return gloss, "definitional"

    # Title-Case technical phrases (2–4 words) recurring in short English.
    titles = re.findall(r"\b([A-Z][a-z]+(?:[-\s][A-Z][a-z]+){0,3})\b", en_clean)
    if titles and len(hi_tokens) <= 10:
        # Prefer multiword titles.
        titles = sorted(titles, key=lambda t: (-len(t.split()), -len(t)))
        for t in titles:
            if t.lower() not in {"the", "this", "that", "when", "every", "human"}:
                return t, "title-case"

    if token in hi_tokens and len(hi_tokens) <= 4:
        gloss = first_sentence(en_clean, max_words=6)
        if gloss:
            return gloss, "compact"

    return None


def pick_evidence(item: dict) -> dict | None:
    hits = [h for h in item.get("evidence", []) if h.get("en")]
    if not hits:
        return None

    def score(h: dict) -> tuple:
        en = clean_english(h["en"])
        hi_n = len(DEVA_TOKEN.findall(h["hi"]))
        en_n = len(LATIN_WORD.findall(en))
        book_rank = 0 if h["book"] == "MVD" else 1
        # Prefer short focused pairs with real English.
        return (book_rank, hi_n + en_n, len(h["hi"]))

    hits.sort(key=score)
    for h in hits:
        if clean_english(h["en"]):
            return h
    return None


def jv_pages_for_english(english: str, jv_text: str) -> list[int]:
    if not english or not jv_text:
        return []
    # Search a distinctive multi-word gloss; skip very short/generic.
    words = LATIN_WORD.findall(english)
    if len(words) < 2:
        return []
    needle = " ".join(words[:4]).lower()
    if len(needle) < 8:
        return []
    pages: list[int] = []
    current = None
    for line in jv_text.splitlines():
        m = re.match(r"^(?:page[-\s]?|\[p\.\s*)(\d+)", line.strip(), re.I)
        if m:
            current = int(m.group(1))
            continue
        if current and needle in line.lower():
            pages.append(current)
            if len(pages) >= 2:
                break
    return pages


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--evidence",
        type=Path,
        default=HERE / "phase4_evidence.json",
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument(
        "--jv-md",
        type=Path,
        default=HERE.parent / "JV-Jeevan-Vidya-An-Introduction.md",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=HERE / "phase4_new_rows.json",
    )
    parser.add_argument(
        "--rejects",
        type=Path,
        default=HERE / "phase4_rejects.json",
    )
    args = parser.parse_args()

    evidence = json.loads(args.evidence.read_text(encoding="utf-8"))
    exact, stems, _lookup = load_known(args.xlsx)
    stop = STOPWORDS | EXTRA_STOP
    jv_text = ""
    if args.jv_md.is_file():
        jv_text = args.jv_md.read_text(encoding="utf-8")

    proposals: list[dict] = []
    rejects: list[dict] = []
    seen_norm: dict[str, int] = {}

    for item in evidence:
        token = item["token"]
        freq = item["total_occurrences"]
        if token in stop or len(token) < 2:
            rejects.append({"token": token, "freq": freq, "reason": "stopword"})
            continue
        if token in exact:
            rejects.append({"token": token, "freq": freq, "reason": "already-in-mapping"})
            continue
        stem = normalize_lemma(token)
        if stem in stems and stem != token:
            # Likely inflection/derivation of an existing lemma.
            rejects.append(
                {
                    "token": token,
                    "freq": freq,
                    "reason": f"inflection-of-known-stem:{stem}",
                }
            )
            continue

        hit = pick_evidence(item)
        if not hit:
            rejects.append({"token": token, "freq": freq, "reason": "no-clean-evidence"})
            continue
        extracted = extract_gloss(token, hit["hi"], hit["en"])
        if not extracted:
            rejects.append({"token": token, "freq": freq, "reason": "no-gloss-extracted"})
            continue
        english, method = extracted
        # Drop ultra-generic English.
        if english.lower() in {
            "the",
            "this",
            "that",
            "it",
            "is",
            "are",
            "human",
            "humans",
            "every",
            "when",
            "and",
            "or",
            "of",
            "in",
            "to",
            "for",
            "with",
            "from",
            "as",
            "by",
            "on",
            "at",
            "a",
            "an",
        }:
            rejects.append({"token": token, "freq": freq, "reason": "generic-english"})
            continue

        # Merge by normalized stem into one lemma row.
        merge_key = stem if stem else token
        if merge_key in seen_norm:
            idx = seen_norm[merge_key]
            row = proposals[idx]
            variants = [v.strip() for v in row["hindi_lemma"].split(",") if v.strip()]
            if token not in variants:
                variants.append(token)
                row["hindi_lemma"] = ", ".join(variants)
            continue

        citation = f"{hit['book']} p.{hit['hi_page']}"
        jv_pages = jv_pages_for_english(english, jv_text)
        if jv_pages:
            citation += "; " + "; ".join(f"JV p.{p}" for p in jv_pages)

        note = (
            f"Phase 4 (freq≥2 exhaustive): English from paired {hit['book']} evidence "
            f"via {method}. Token freq={freq} in MVD/SB pairs."
        )
        seen_norm[merge_key] = len(proposals)
        proposals.append(
            {
                "token": token,
                "action": "propose",
                "hindi_lemma": token,
                "english": english,
                "transliteration": transliterate(token),
                "note": note,
                "citation": citation,
                "freq": freq,
                "method": method,
                "evidence_book": hit["book"],
                "evidence_page": hit["hi_page"],
            }
        )

    args.output.write_text(
        json.dumps(proposals, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    args.rejects.write_text(
        json.dumps(rejects, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    by_reason: dict[str, int] = {}
    for r in rejects:
        by_reason[r["reason"].split(":")[0]] = by_reason.get(r["reason"].split(":")[0], 0) + 1
    print(
        f"proposals={len(proposals)} rejects={len(rejects)} reasons={by_reason}",
        file=sys.stderr,
    )
    print(f"wrote {args.output}", file=sys.stderr)
    print(f"wrote {args.rejects}", file=sys.stderr)


if __name__ == "__main__":
    main()
