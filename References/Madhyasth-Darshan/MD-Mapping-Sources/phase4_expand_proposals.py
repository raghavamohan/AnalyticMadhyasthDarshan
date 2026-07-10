"""Expand phase4_new_rows with single-token units and quote/title glosses."""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from phase4_propose_rows import (
    EXTRA_STOP,
    clean_english,
    first_sentence,
    normalize_lemma,
    transliterate,
)
from stopwords import STOPWORDS

HERE = Path(__file__).resolve().parent
XLSX = HERE.parent / "MD-Mapping.xlsx"
DEVA = re.compile(r"[ऀ-ॿ]+")
LATIN = re.compile(r"[A-Za-z][A-Za-z'\-/]*")
QUOTED = re.compile(r"[\"'‘’“”]([^\"'‘’“”]{2,40})[\"'‘’“”]")
BAD = {
    "जब",
    "तब",
    "बात",
    "ने",
    "जा",
    "सब",
    "आज",
    "अत",
    "यातु",
    "कम",
    "बन",
    "नाम",
    "अध्याय",
    "दो",
    "तीन",
    "चार",
    "पाँच",
    "पांच",
    "एक",
    "नागराज",
    "सफल",
    "सुलभ",
    "संभव",
    "प्रस्तुत",
    "समय",
    "देश",
    "स्थान",
    "दिशा",
    "भाषा",
    "वर्ग",
    "कुल",
    "अंग",
}


def load_known() -> tuple[set[str], set[str]]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    exact: set[str] = set()
    stems: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        for part in str(row[0]).replace("،", ",").split(","):
            cell = part.strip()
            if not cell:
                continue
            exact.add(cell)
            exact.update(DEVA.findall(cell))
            stems.add(normalize_lemma(cell))
            for tok in DEVA.findall(cell):
                stems.add(normalize_lemma(tok))
    wb.close()
    return exact, stems


def main() -> None:
    exact, stems = load_known()
    freq = {
        t: f
        for t, f in json.loads((HERE / "candidates_freq2.json").read_text(encoding="utf-8"))
    }
    existing = json.loads((HERE / "phase4_new_rows.json").read_text(encoding="utf-8"))
    have = {r["token"] for r in existing}
    stop = STOPWORDS | EXTRA_STOP | BAD
    added: list[dict] = []

    for name in ("mvd_pairs.json", "sb_pairs.json"):
        for p in json.loads((HERE / name).read_text(encoding="utf-8")):
            if not p.get("en"):
                continue
            toks = DEVA.findall(p["hi"])
            if len(toks) != 1:
                continue
            tok = toks[0]
            if tok in have or tok in exact or tok in stop or freq.get(tok, 0) < 2:
                continue
            stem = normalize_lemma(tok)
            if stem in stems and stem != tok:
                continue
            gloss = first_sentence(clean_english(p["en"]), max_words=5)
            if not gloss:
                continue
            words = LATIN.findall(gloss)
            if not (1 <= len(words) <= 5):
                continue
            low = gloss.lower()
            if low.startswith(("chapter", "may ", "page", "http", "according", "every ", "when ")):
                continue
            added.append(
                {
                    "token": tok,
                    "action": "propose",
                    "hindi_lemma": tok,
                    "english": gloss,
                    "transliteration": transliterate(tok),
                    "note": (
                        f"Phase 4: single-token bilingual unit ({p['book']}). "
                        f"freq={freq.get(tok)}."
                    ),
                    "citation": f"{p['book']} p.{p['hi_page']}",
                    "freq": freq.get(tok),
                    "method": "single-token-hi",
                }
            )
            have.add(tok)

    evidence = json.loads((HERE / "phase4_evidence.json").read_text(encoding="utf-8"))
    for item in evidence:
        tok = item["token"]
        if tok in have or tok in exact or tok in stop or item["total_occurrences"] < 2:
            continue
        stem = normalize_lemma(tok)
        if stem in stems and stem != tok:
            continue
        for h in item["evidence"]:
            en = clean_english(h.get("en") or "")
            if not en:
                continue
            hi_toks = DEVA.findall(h["hi"])
            if tok not in hi_toks or len(hi_toks) > 10:
                continue
            gloss = None
            for q in QUOTED.findall(en):
                w = LATIN.findall(q)
                if 1 <= len(w) <= 4:
                    gloss = q.strip()
                    break
            if not gloss:
                titles = re.findall(r"\b([A-Z][a-z]+(?:[-\s][A-Z][a-z]+){1,3})\b", en)
                if titles:
                    cand = sorted(titles, key=lambda t: -len(t))[0]
                    if 2 <= len(LATIN.findall(cand)) <= 4:
                        gloss = cand
            if not gloss:
                continue
            if gloss.lower() in {"jeevan", "big", "world", "mirage", "kindness", "conceit"}:
                continue
            added.append(
                {
                    "token": tok,
                    "action": "propose",
                    "hindi_lemma": tok,
                    "english": gloss,
                    "transliteration": transliterate(tok),
                    "note": (
                        f"Phase 4: quoted/title gloss from focused evidence ({h['book']}). "
                        f"freq={item['total_occurrences']}."
                    ),
                    "citation": f"{h['book']} p.{h['hi_page']}",
                    "freq": item["total_occurrences"],
                    "method": "quote-or-title",
                }
            )
            have.add(tok)
            break

    out: list[dict] = []
    seen: set[str] = set()
    for row in existing + added:
        if row["token"] in seen:
            continue
        seen.add(row["token"])
        out.append(row)
    out.sort(key=lambda r: (-(r.get("freq") or 0), r["hindi_lemma"]))
    (HERE / "phase4_new_rows.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=1) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(
        f"added={len(added)} total={len(out)} methods={Counter(r['method'] for r in out)}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
